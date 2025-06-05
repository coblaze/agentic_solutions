"""
Excel Report Generation Service
Creates formatted Excel reports with metrics and detailed evaluation data
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging
from src.models.evaluation import EvaluationResult, BatchEvaluation
from src.config.settings import settings

logger = logging.getLogger(__name__)


class ReportService:
    """
    Service for generating Excel evaluation reports
    Creates professional reports with metrics and detailed data
    """
    
    def __init__(self):
        self.output_dir = os.path.join(os.getcwd(), "evaluation_reports")
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Define styles
        self._init_styles()
        
    def _init_styles(self):
        """Initialize Excel styles for consistent formatting"""
        # Colors
        self.header_color = "366092"
        self.subheader_color = "4472C4"
        self.pass_color = "70AD47"
        self.fail_color = "FF0000"
        self.warning_color = "FFC000"
        
        # Fonts
        self.title_font = Font(name='Calibri', size=16, bold=True)
        self.header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        self.subheader_font = Font(name='Calibri', size=11, bold=True)
        self.normal_font = Font(name='Calibri', size=11)
        self.bold_font = Font(name='Calibri', size=11, bold=True)
        
        # Fills
        self.header_fill = PatternFill(start_color=self.header_color, 
                                      end_color=self.header_color, 
                                      fill_type="solid")
        self.subheader_fill = PatternFill(start_color=self.subheader_color, 
                                         end_color=self.subheader_color, 
                                         fill_type="solid")
        self.pass_fill = PatternFill(start_color=self.pass_color, 
                                    end_color=self.pass_color, 
                                    fill_type="solid")
        self.fail_fill = PatternFill(start_color=self.fail_color, 
                                    end_color=self.fail_color, 
                                    fill_type="solid")
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignment
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
    def generate_excel_report(self, 
                            batch: BatchEvaluation, 
                            results: List[EvaluationResult],
                            password: Optional[str] = None) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            batch: Batch evaluation summary
            results: List of evaluation results
            password: Optional password for Excel file (not implemented in openpyxl)
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Generate filename
            date_str = batch.evaluation_date.strftime("%m_%d_%Y")
            filename = f"{date_str}_Evaluation_Report.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                
            # Create sheets
            self._create_metrics_sheet(wb, batch)
            self._create_evaluation_data_sheet(wb, results)
            self._create_summary_stats_sheet(wb, batch, results)
            
            # Add document properties
            wb.properties.title = f"Evaluation Report - {date_str}"
            wb.properties.creator = "LLM Evaluation Pipeline"
            wb.properties.description = f"Automated evaluation report for {batch.evaluation_date.strftime('%Y-%m-%d')}"
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Excel report generated: {filepath}")
            
            # Note: openpyxl doesn't support password protection directly
            # For production, consider using win32com (Windows) or other libraries
            if password:
                logger.warning("Password protection requested but not implemented in openpyxl")
                
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
            
    def _create_metrics_sheet(self, wb: Workbook, batch: BatchEvaluation):
        """Create metrics summary sheet"""
        ws = wb.create_sheet("Metrics")
        
        # Title
        ws['A1'] = "Evaluation Metrics Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Evaluation info
        ws['A3'] = "Report Details"
        ws['A3'].font = self.subheader_font
        
        info_data = [
            ("Evaluation Date:", batch.evaluation_date.strftime("%Y-%m-%d %H:%M:%S")),
            ("Batch ID:", batch.batch_id),
            ("Report Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")),
            ("Processing Time:", f"{batch.processing_time_seconds:.1f} seconds" if batch.processing_time_seconds else "N/A")
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.bold_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.normal_font
            row += 1
            
        # Metrics section
        ws[f'A{row + 1}'] = "Evaluation Metrics"
        ws[f'A{row + 1}'].font = self.subheader_font
        
        # Create metrics table
        metrics_headers = ["Metric", "Value", "Status"]
        metrics_data = [
            ("Total Evaluations", batch.total_evaluations, "info"),
            ("Passed", f"{batch.passed} ({batch.pass_percentage:.2f}%)", "pass"),
            ("Failed", f"{batch.failed} ({batch.fail_percentage:.2f}%)", "fail"),
            ("Errors", f"{batch.errors} ({batch.error_percentage:.2f}%)" if batch.errors > 0 else "0 (0.00%)", "error"),
            ("Overall Accuracy", f"{batch.accuracy:.2%}", "pass" if batch.accuracy >= settings.ACCURACY_THRESHOLD else "fail"),
            ("Accuracy Threshold", f"{settings.ACCURACY_THRESHOLD:.0%}", "info"),
            ("Average Confidence", f"{batch.average_confidence:.2%}" if batch.average_confidence else "N/A", "info")
        ]
        
        # Write headers
        start_row = row + 3
        for col, header in enumerate(metrics_headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
        # Write data
        for i, (metric, value, status) in enumerate(metrics_data, 1):
            row_num = start_row + i
            
            # Metric name
            ws.cell(row=row_num, column=1, value=metric).font = self.bold_font
            
            # Value
            value_cell = ws.cell(row=row_num, column=2, value=value)
            value_cell.font = self.normal_font
            
            # Status coloring
            status_cell = ws.cell(row=row_num, column=3)
            if status == "pass":
                status_cell.value = "✓"
                status_cell.font = Font(color=self.pass_color, bold=True)
            elif status == "fail":
                status_cell.value = "✗"
                status_cell.font = Font(color=self.fail_color, bold=True)
            elif status == "error":
                status_cell.value = "⚠"
                status_cell.font = Font(color=self.warning_color, bold=True)
            else:
                status_cell.value = "•"
                
            # Apply borders
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).border = self.thin_border
                
        # Add alert status
        alert_row = start_row + len(metrics_data) + 2
        ws[f'A{alert_row}'] = "Alert Status:"
        ws[f'A{alert_row}'].font = self.bold_font
        
        if batch.accuracy < settings.ACCURACY_THRESHOLD:
            ws[f'B{alert_row}'] = "ALERT TRIGGERED - Accuracy below threshold"
            ws[f'B{alert_row}'].font = Font(color=self.fail_color, bold=True)
        else:
            ws[f'B{alert_row}'] = "No alerts - Accuracy within acceptable range"
            ws[f'B{alert_row}'].font = Font(color=self.pass_color, bold=True)
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
    def _create_evaluation_data_sheet(self, wb: Workbook, results: List[EvaluationResult]):
        """Create detailed evaluation data sheet"""
        ws = wb.create_sheet("Evaluation Data")
        
        # Convert results to dataframe
        data = []
        for result in results:
            data.append(result.to_report_dict())
            
        df = pd.DataFrame(data)
        
        # Write dataframe to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_align
                    cell.border = self.thin_border
                else:
                    cell.font = self.normal_font
                    cell.border = self.thin_border
                    
                    # Color code status column
                    if c_idx == df.columns.get_loc('evaluationStatus') + 1:
                        if value == 'pass':
                            cell.font = Font(color=self.pass_color, bold=True)
                        elif value == 'fail':
                            cell.font = Font(color=self.fail_color, bold=True)
                            
                    # Wrap text for long columns
                    if c_idx in [df.columns.get_loc('message') + 1, 
                               df.columns.get_loc('postCallSummary') + 1,
                               df.columns.get_loc('evaluationReason') + 1]:
                        cell.alignment = self.wrap_align
                        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Limit max width for readability
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
                    
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Set specific widths for text-heavy columns
        ws.column_dimensions['G'].width = 60  # message
        ws.column_dimensions['I'].width = 40  # postCallSummary
        ws.column_dimensions['L'].width = 50  # evaluationReason
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
    def _create_summary_stats_sheet(self, wb: Workbook, batch: BatchEvaluation, results: List[EvaluationResult]):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary Statistics")
        
        # Title
        ws['A1'] = "Summary Statistics"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        # Group results by various dimensions
        df = pd.DataFrame([r.to_report_dict() for r in results])
        
        # By LOB
        ws['A3'] = "Performance by Line of Business"
        ws['A3'].font = self.subheader_font
        
        if not df.empty and 'LOB' in df.columns:
            lob_stats = df.groupby('LOB')['evaluationStatus'].value_counts().unstack(fill_value=0)
            
            row = 5
            # Headers
            ws.cell(row=row, column=1, value="LOB").font = self.header_font
            ws.cell(row=row, column=2, value="Passed").font = self.header_font
            ws.cell(row=row, column=3, value="Failed").font = self.header_font
            ws.cell(row=row, column=4, value="Accuracy").font = self.header_font
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = self.header_fill
                ws.cell(row=row, column=col).alignment = self.center_align
                
            # Data
            for lob in lob_stats.index:
                row += 1
                passed = lob_stats.loc[lob].get('pass', 0)
                failed = lob_stats.loc[lob].get('fail', 0)
                total = passed + failed
                accuracy = (passed / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=1, value=lob)
                ws.cell(row=row, column=2, value=passed)
                ws.cell(row=row, column=3, value=failed)
                ws.cell(row=row, column=4, value=f"{accuracy:.2f}%")
                
                # Color code accuracy
                acc_cell = ws.cell(row=row, column=4)
                if accuracy >= 85:
                    acc_cell.font = Font(color=self.pass_color, bold=True)
                else:
                    acc_cell.font = Font(color=self.fail_color, bold=True)
                    
        # By Hour (if applicable)
        if not df.empty and 'startTimestamp' in df.columns:
            ws[f'A{row + 3}'] = "Performance by Hour"
            ws[f'A{row + 3}'].font = self.subheader_font
            
            # Parse timestamps and extract hour
            df['hour'] = pd.to_datetime(df['startTimestamp']).dt.hour
            hourly_stats = df.groupby('hour')['evaluationStatus'].value_counts().unstack(fill_value=0)
            
            row += 5
            # Headers
            ws.cell(row=row, column=1, value="Hour").font = self.header_font
            ws.cell(row=row, column=2, value="Total Calls").font = self.header_font
            ws.cell(row=row, column=3, value="Accuracy").font = self.header_font
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = self.header_fill
                ws.cell(row=row, column=col).alignment = self.center_align
                
            # Data
            for hour in sorted(hourly_stats.index):
                row += 1
                passed = hourly_stats.loc[hour].get('pass', 0)
                failed = hourly_stats.loc[hour].get('fail', 0)
                total = passed + failed
                accuracy = (passed / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=1, value=f"{hour:02d}:00")
                ws.cell(row=row, column=2, value=total)
                ws.cell(row=row, column=3, value=f"{accuracy:.2f}%")
                
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20